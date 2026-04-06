import streamlit as st
import tempfile
import os
import io
import re
from collections import Counter

import pandas as pd
import ifcopenshell
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from engine.semantic.group_phrases import group_elements
from engine.semantic.embedding_classifier import EmbeddingClassifier
from engine.semantic.semantic_translator import build_phrase_ss, build_phrase_pr
from engine.rules.domain_classifier import detect_domain_from_ifc
from engine.rules.ef_classifier import get_ef_candidates
from engine.rules.ss_classifier import classify_ss
from engine.rules.pr_classifier import classify_pr

# -----------------------------------------------
# PAGE CONFIG
# -----------------------------------------------

st.set_page_config(
    page_title="UNIBIM Uniclass Classifier",
    page_icon="🏗️",
    layout="wide",
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #1e3a5f 0%, #2d6a9f 100%);
        padding: 1.5rem 2rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .main-header h1 { margin: 0; font-size: 2rem; color: white; }
    .main-header p  { margin: 0.3rem 0 0; opacity: 0.85; font-size: 0.95rem; }

    .metric-card {
        background: #f8f9fa;
        border-radius: 10px;
        padding: 1rem 1.2rem;
        border-left: 4px solid #2d6a9f;
        text-align: center;
    }
    .metric-ok     { border-left-color: #28a745; }
    .metric-warn   { border-left-color: #ffc107; }
    .metric-total  { border-left-color: #2d6a9f; }

    .conf-bar-wrap { width: 100%; background: #e9ecef; border-radius: 4px; height: 10px; }
    .conf-bar      { height: 10px; border-radius: 4px; }

    [data-testid="stSidebar"] { background: #f0f4f8; }
    [data-testid="stSidebar"] h2 { color: #1e3a5f; }

    div[data-testid="stTabs"] button { font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# -----------------------------------------------
# CONSTANTS
# -----------------------------------------------

DISCIPLINE_MAP = {
    "ARQ": "architecture",
    "HVAC": "hvac",
    "PCI": "fire",
    "SAN": "sanitary",
    "EST": "structural",
    "ELE": "electrical",
}

DISCIPLINE_EMOJI = {
    "architecture": "🏛️",
    "hvac": "🌀",
    "fire": "🔥",
    "sanitary": "💧",
    "structural": "🏗️",
    "electrical": "⚡",
}

CONFIDENCE_THRESHOLD = 0.75

IFC_SKIP = {
    "ifcopeningelement", "ifcspace",
    "ifcvirtualelement", "ifcannotation", "ifcelementassembly",
}

# -----------------------------------------------
# RESOURCE LOADING
# -----------------------------------------------

@st.cache_resource(show_spinner="Cargando clasificadores...")
def load_resources():
    ef_text, ss_text, pr_text = {}, {}, {}

    uniclass_dir = "data/uniclass"
    if os.path.isdir(uniclass_dir):
        try:
            wb = openpyxl.load_workbook(f"{uniclass_dir}/Uniclass2015_EF_v1_16.xlsx", read_only=True)
            for row in wb.active.iter_rows(values_only=True):
                code, title = str(row[0] or "").strip(), str(row[6] or "").strip()
                if code.startswith("EF_") and title:
                    ef_text[code.lower()] = title
            wb.close()
        except Exception:
            pass

        try:
            wb = openpyxl.load_workbook(f"{uniclass_dir}/Uniclass2015_Ss_v1_40.xlsx", read_only=True)
            for row in wb.active.iter_rows(min_row=4, values_only=True):
                code, title = str(row[0] or "").strip(), str(row[6] or "").strip()
                if code.startswith("Ss_") and title:
                    ss_text[code.lower()] = title
            wb.close()
        except Exception:
            pass

        try:
            wb = openpyxl.load_workbook(f"{uniclass_dir}/Uniclass2015_Pr_v1_40.xlsx", read_only=True)
            for row in wb.active.iter_rows(min_row=4, values_only=True):
                code, title = str(row[0] or "").strip(), str(row[6] or "").strip()
                if code.startswith("Pr_") and title:
                    pr_text[code.lower()] = title
            wb.close()
        except Exception:
            pass

    classifier_ss = EmbeddingClassifier(
        model_path="models/construction_embedding_model",
        vectors_path="data/processed/ss_vectors.npy",
        metadata_path="data/processed/ss_metadata.json",
    )
    classifier_pr = EmbeddingClassifier(
        model_path="models/construction_embedding_model",
        vectors_path="data/processed/pr_vectors.npy",
        metadata_path="data/processed/pr_metadata.json",
    )

    return ef_text, ss_text, pr_text, classifier_ss, classifier_pr


# -----------------------------------------------
# HELPERS
# -----------------------------------------------

def extract_discipline(filename):
    match = re.search(r"IFC-([A-Z]+)-\d+", filename, re.IGNORECASE)
    return match.group(1).upper() if match else ""


def clean_name(raw):
    if not raw or raw in ("$", ""):
        return None
    parts = raw.rsplit(":", 1)
    if len(parts) == 2 and parts[1].strip().isdigit():
        raw = parts[0].strip()
    return raw or None


def should_skip(data, ef_candidates):
    if not ef_candidates or ef_candidates == ["ef_00"]:
        return True
    return data.get("ifc_type", "").lower() in IFC_SKIP


def conf_color(pct_str):
    """Return a hex color based on confidence percentage string like '82%'."""
    try:
        val = int(pct_str.replace("%", "")) / 100
    except Exception:
        return "#adb5bd"
    if val >= 0.85:
        return "#28a745"
    if val >= 0.75:
        return "#20c997"
    if val >= 0.60:
        return "#ffc107"
    return "#dc3545"


def conf_bar_html(pct_str):
    color = conf_color(pct_str)
    try:
        val = int(pct_str.replace("%", ""))
    except Exception:
        val = 0
    return (
        f'<div class="conf-bar-wrap">'
        f'<div class="conf-bar" style="width:{val}%;background:{color};"></div>'
        f'</div>'
        f'<small style="color:{color};font-weight:600">{pct_str}</small>'
    )


# -----------------------------------------------
# CLASSIFICATION CORE
# -----------------------------------------------

def classify_ifc_file(ifc_path, model_name, ef_text, ss_text, pr_text,
                      classifier_ss, classifier_pr):
    try:
        model = ifcopenshell.open(ifc_path)
    except Exception as e:
        st.error(f"No se pudo abrir {model_name}: {e}")
        return []

    discipline = DISCIPLINE_MAP.get(model_name, "")
    elements   = model.by_type("IfcElement")
    domain     = detect_domain_from_ifc(elements)
    groups     = group_elements(elements)

    preprocessed = []
    for phrase_key, items in groups.items():
        data = dict(items[0])
        data["discipline"] = discipline
        ef_candidates = get_ef_candidates(data, domain)
        if should_skip(data, ef_candidates):
            continue
        ef_main = ef_candidates[0]
        preprocessed.append({
            "phrase_key": phrase_key,
            "items":      items,
            "data":       data,
            "ef_main":    ef_main,
            "phrase_ss":  build_phrase_ss(data),
            "phrase_pr":  build_phrase_pr(data),
        })

    classifier_ss.encode_batch([g["phrase_ss"] for g in preprocessed])
    classifier_pr.encode_batch([g["phrase_pr"] for g in preprocessed])

    results = []
    for g in preprocessed:
        data    = g["data"]
        ef_main = g["ef_main"]

        ss_raw   = classifier_ss.classify(g["phrase_ss"], top_k=3)
        ss_final = classify_ss(data, ss_raw)

        pr_raw   = classifier_pr.classify(g["phrase_pr"], top_k=5, ef=ef_main, data=data)
        pr_final = classify_pr(data, pr_raw)

        ss_code = ss_final["code"]
        ss_text_val = ss_text.get(ss_code, ss_final.get("text", ""))

        pr_code = pr_final["code"]
        pr_text_val = pr_text.get(pr_code, pr_final.get("text", ""))

        ef_text_val = ef_text.get(ef_main, "")

        ss_conf = ss_final.get("confidence", 0.0)
        pr_conf = pr_final.get("confidence", 0.0)
        overall = min(ss_conf, pr_conf)

        name_counts = Counter(clean_name(i.get("name", "$")) for i in g["items"])
        name_counts.pop(None, None)
        element_names = " | ".join(
            f"{n} ({c})" for n, c in sorted(name_counts.items(), key=lambda x: -x[1])
        )

        results.append({
            "Modelo":         model_name,
            "Disciplina":     DISCIPLINE_EMOJI.get(discipline, "") + " " + discipline.upper() if discipline else "—",
            "Dominio":        domain,
            "Frase":          g["phrase_key"],
            "Elementos":      len(g["items"]),
            "Nombres":        element_names,
            "EF":             ef_main.upper(),
            "EF Descripción": ef_text_val,
            "Ss":             ss_code.upper(),
            "Ss Descripción": ss_text_val,
            "Ss Fuente":      ss_final.get("source", ""),
            "Pr":             pr_code.upper(),
            "Pr Descripción": pr_text_val,
            "Pr Fuente":      pr_final.get("source", ""),
            "_conf_val":      overall,
            "Confianza":      f"{overall:.0%}",
            "Estado":         "✅ OK" if overall >= CONFIDENCE_THRESHOLD else "⚠️ Revisar",
        })

    return results


# -----------------------------------------------
# EXCEL EXPORT
# -----------------------------------------------

def build_excel(df: pd.DataFrame) -> bytes:
    """Return a formatted Excel workbook as bytes."""

    export_cols = [
        "Modelo", "Disciplina", "Dominio", "Frase", "Elementos", "Nombres",
        "EF", "EF Descripción",
        "Ss", "Ss Descripción", "Ss Fuente",
        "Pr", "Pr Descripción", "Pr Fuente",
        "Confianza", "Estado",
    ]
    # keep only cols that exist
    export_cols = [c for c in export_cols if c in df.columns]
    out = df[export_cols].copy()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clasificación BIM"

    # ── Styles ──────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    OK_FILL    = PatternFill("solid", fgColor="D6F5E3")   # light green
    WARN_FILL  = PatternFill("solid", fgColor="FFF8E1")   # light amber
    ALT_FILL   = PatternFill("solid", fgColor="F0F4F8")   # alternate row
    BASE_FONT  = Font(size=9, name="Calibri")
    WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
    CENTER     = Alignment(horizontal="center", vertical="top")

    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ───────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_cols))
    title_cell = ws.cell(row=1, column=1,
                         value="UNIBIM Uniclass Classifier · Uniclass 2015")
    title_cell.font      = Font(bold=True, size=13, color="1E3A5F", name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill      = PatternFill("solid", fgColor="E8EEF5")
    ws.row_dimensions[1].height = 28

    # ── Sub-title row ────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(export_cols))
    sub_cell = ws.cell(row=2, column=1,
                       value="Integrantes: Alejandro Martínez · Camilo Torres · Gissela Chicaiza · "
                             "Maite Castiñeira · Pablo Pinuer · Santiago Martínez Chabbert   |   "
                             "Tutor: Evelio Sanchez   |   Zigurat Institute · © 2026 UNIBIM")
    sub_cell.font      = Font(italic=True, size=8, color="555555", name="Calibri")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    sub_cell.fill      = PatternFill("solid", fgColor="F5F7FA")
    ws.row_dimensions[2].height = 18

    # ── Header row (row 3) ───────────────────────────────────
    HEADER_LABELS = {
        "Modelo":          "Modelo",
        "Disciplina":      "Disciplina",
        "Dominio":         "Dominio",
        "Frase":           "Frase / Grupo",
        "Elementos":       "Cant.\nElementos",
        "Nombres":         "Nombres de elementos",
        "EF":              "EF\nCódigo",
        "EF Descripción":  "EF — Descripción",
        "Ss":              "Ss\nCódigo",
        "Ss Descripción":  "Ss — Descripción",
        "Ss Fuente":       "Ss\nFuente",
        "Pr":              "Pr\nCódigo",
        "Pr Descripción":  "Pr — Descripción",
        "Pr Fuente":       "Pr\nFuente",
        "Confianza":       "Confianza",
        "Estado":          "Estado",
    }
    for col_idx, col_name in enumerate(export_cols, start=1):
        cell = ws.cell(row=3, column=col_idx, value=HEADER_LABELS.get(col_name, col_name))
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border    = BORDER
    ws.row_dimensions[3].height = 32

    # ── Data rows ────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(out.iterrows(), start=4):
        estado = str(row.get("Estado", ""))
        if estado == "✅ OK":
            row_fill = OK_FILL
        elif "Revisar" in estado:
            row_fill = WARN_FILL
        else:
            row_fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(export_cols, start=1):
            val  = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = BASE_FONT
            cell.border = BORDER

            if col_name in ("Elementos",):
                cell.alignment = CENTER
            elif col_name in ("EF", "Ss", "Pr", "Ss Fuente", "Pr Fuente",
                              "Confianza", "Estado", "Modelo", "Disciplina", "Dominio"):
                cell.alignment = CENTER
            else:
                cell.alignment = WRAP_ALIGN

            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 40

    # ── Column widths ─────────────────────────────────────────
    COL_WIDTHS = {
        "Modelo": 8, "Disciplina": 14, "Dominio": 12, "Frase": 30,
        "Elementos": 9, "Nombres": 35,
        "EF": 10, "EF Descripción": 32,
        "Ss": 14, "Ss Descripción": 36, "Ss Fuente": 10,
        "Pr": 14, "Pr Descripción": 36, "Pr Fuente": 10,
        "Confianza": 10, "Estado": 10,
    }
    for col_idx, col_name in enumerate(export_cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 14)

    # ── Freeze panes below header ──────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter on header ──────────────────────────────
    ws.auto_filter.ref = (
        f"A3:{get_column_letter(len(export_cols))}3"
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -----------------------------------------------
# HEADER
# -----------------------------------------------

st.markdown("""
<div class="main-header">
    <h1> UNIBIM Uniclass Classifier</h1>
    <p>Clasificación automática Uniclass 2015 &nbsp;·&nbsp; EF &nbsp;·&nbsp; Ss &nbsp;·&nbsp; Pr &nbsp;·&nbsp; desde modelos IFC</p>
</div>
""", unsafe_allow_html=True)

# -----------------------------------------------
# SIDEBAR
# -----------------------------------------------

with st.sidebar:
    # Logo + institución
    st.image("UNIBIM LOGO.jpeg", use_container_width=True)
    st.markdown(
        "<div style='text-align:center;font-size:0.8rem;color:#1e3a5f;font-weight:600;margin-top:4px'>"
        "Zigurat Institute · 2026"
        "</div>",
        unsafe_allow_html=True,
    )
    st.divider()

    st.markdown("## ⚙️ Configuración")
    st.divider()

    uploaded = st.file_uploader(
        "📂 Modelos IFC",
        type=["ifc"],
        accept_multiple_files=True,
        help="Puedes subir modelos de distintas disciplinas (EST, HVAC, PCI, SAN, ELE, ARQ)",
    )

    if uploaded:
        st.markdown("**Archivos cargados:**")
        for f in uploaded:
            disc = extract_discipline(f.name)
            label = DISCIPLINE_MAP.get(disc, "desconocida")
            emoji = DISCIPLINE_EMOJI.get(label, "📄")
            st.markdown(f"{emoji} `{f.name.split('.')[0]}`  \n<small>{label.upper()}</small>", unsafe_allow_html=True)

    st.divider()
    run_btn = st.button("▶  Clasificar", type="primary", disabled=not uploaded, use_container_width=True)

    st.divider()
    st.markdown("### 🔍 Filtros")
    estado_filter = st.selectbox("Estado", ["Todos", "✅ OK", "⚠️ Revisar"])

    if "results_df" in st.session_state:
        df_all = st.session_state["results_df"]
        modelo_opts = sorted(df_all["Modelo"].unique())
        modelo_filter = st.multiselect("Modelo", modelo_opts, default=modelo_opts)
        dominio_opts = sorted(df_all["Dominio"].unique())
        dominio_filter = st.multiselect("Dominio", dominio_opts, default=dominio_opts)
    else:
        modelo_filter  = []
        dominio_filter = []

    st.divider()

    # Créditos
    st.markdown("""
<div style='font-size:0.75rem;color:#444;line-height:1.7'>
<b style='color:#1e3a5f'>Integrantes</b><br>
Alejandro Martínez<br>
Camilo Torres<br>
Gissela Chicaiza<br>
Maite Castiñeira<br>
Pablo Pinuer<br>
Santiago Martínez Chabbert<br>
<br>
<b style='color:#1e3a5f'>Tutor</b><br>
Evelio Sanchez<br>
<br>
<span style='font-size:0.7rem;color:#888'>
© 2026 UNIBIM<br>
UNIBIM Uniclass Classifier<br>
Uniclass 2015 · IFC 4
</span>
</div>
""", unsafe_allow_html=True)


# -----------------------------------------------
# CLASSIFICATION RUN
# -----------------------------------------------

if run_btn:
    ef_text, ss_text, pr_text, classifier_ss, classifier_pr = load_resources()

    all_results = []
    progress    = st.progress(0, text="Iniciando...")
    status_box  = st.empty()
    tmpdir      = tempfile.mkdtemp()

    for i, uploaded_file in enumerate(uploaded):
        model_name = extract_discipline(uploaded_file.name) or uploaded_file.name.split(".")[0].upper()
        status_box.info(f"Procesando **{uploaded_file.name}** ({i+1}/{len(uploaded)})...")

        tmp_path = os.path.join(tmpdir, uploaded_file.name)
        with open(tmp_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        results = classify_ifc_file(
            tmp_path, model_name,
            ef_text, ss_text, pr_text,
            classifier_ss, classifier_pr,
        )
        all_results.extend(results)
        os.unlink(tmp_path)
        progress.progress((i + 1) / len(uploaded), text=f"{uploaded_file.name} listo")

    os.rmdir(tmpdir)
    progress.empty()
    status_box.empty()

    if not all_results:
        st.warning("No se clasificaron elementos. Verifica que los archivos IFC sean válidos.")
        st.stop()

    st.session_state["results_df"] = pd.DataFrame(all_results)


# -----------------------------------------------
# TOP-LEVEL TABS
# -----------------------------------------------

tab_app, tab_flow = st.tabs(["📊  Clasificación", "🔄  Metodología"])

# -----------------------------------------------
# TAB: CLASIFICACIÓN
# -----------------------------------------------

with tab_app:
    if "results_df" not in st.session_state:
        st.info("Sube uno o más archivos IFC desde el panel lateral y pulsa **▶ Clasificar**.")
    else:
        df = st.session_state["results_df"]

        # Apply sidebar filters
        view = df.copy()
        if estado_filter != "Todos":
            view = view[view["Estado"] == estado_filter]
        if modelo_filter:
            view = view[view["Modelo"].isin(modelo_filter)]
        if dominio_filter:
            view = view[view["Dominio"].isin(dominio_filter)]

        # KPI row
        total    = len(df)
        ok       = (df["Estado"] == "✅ OK").sum()
        revisar  = (df["Estado"] == "⚠️ Revisar").sum()
        modelos  = df["Modelo"].nunique()
        avg_conf = df["_conf_val"].mean()

        k1, k2, k3, k4, k5, k6 = st.columns([1, 1, 1, 1, 1, 1.4])
        k1.metric("Grupos clasificados", total)
        k2.metric("✅ OK",               ok,      f"{ok/total:.0%}")
        k3.metric("⚠️ Revisar",          revisar, f"{revisar/total:.0%}", delta_color="inverse")
        k4.metric("Modelos procesados",  modelos)
        k5.metric("Confianza promedio",  f"{avg_conf:.0%}")
        with k6:
            st.markdown("<div style='height:0.6rem'></div>", unsafe_allow_html=True)
            excel_bytes_top = build_excel(df)
            st.download_button(
                label="⬇️  Descargar",
                data=excel_bytes_top,
                file_name="clasificacion_bim.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        st.divider()

        # ---- SUB-TABS ----
        sub_resumen, sub_detalle = st.tabs(["📈  Resumen", "📋  Detalle"])

        # ---- SUB-TAB: RESUMEN ----
        with sub_resumen:
            col_left, col_right = st.columns(2)

            with col_left:
                st.markdown("#### Estado de clasificación")
                pie_data = pd.DataFrame({
                    "Estado": ["✅ OK", "⚠️ Revisar"],
                    "Cantidad": [ok, revisar],
                })
                st.bar_chart(pie_data.set_index("Estado"), color=["#28a745"], height=260)

            with col_right:
                st.markdown("#### Grupos por dominio")
                domain_counts = df.groupby("Dominio").size().reset_index(name="Grupos")
                st.bar_chart(domain_counts.set_index("Dominio"), height=260)

            st.divider()
            col_a, col_b = st.columns(2)

            with col_a:
                st.markdown("#### Confianza por modelo")
                conf_model = (
                    df.groupby("Modelo")["_conf_val"]
                    .mean()
                    .reset_index()
                    .rename(columns={"_conf_val": "Confianza promedio"})
                    .sort_values("Confianza promedio", ascending=False)
                )
                conf_model["Confianza promedio"] = (conf_model["Confianza promedio"] * 100).round(1)
                st.dataframe(conf_model, width="stretch", hide_index=True)

            with col_b:
                st.markdown("#### Top 10 — Ss más frecuentes")
                ss_counts = (
                    df.groupby(["Ss", "Ss Descripción"])
                    .size()
                    .reset_index(name="Grupos")
                    .sort_values("Grupos", ascending=False)
                    .head(10)
                )
                st.dataframe(ss_counts, width="stretch", hide_index=True)

        # ---- SUB-TAB: DETALLE ----
        with sub_detalle:
            display_cols = [
                "Modelo", "Dominio", "Frase", "Elementos",
                "EF", "EF Descripción",
                "Ss", "Ss Descripción",
                "Pr", "Pr Descripción",
                "Confianza", "Estado",
            ]

            def highlight_estado(row):
                if row["Estado"] == "⚠️ Revisar":
                    return ["background-color: #fff8e1; color: #5c4300"] * len(row)
                return [""] * len(row)

            styled = view[display_cols].style.apply(highlight_estado, axis=1)
            st.dataframe(styled, width="stretch", height=520)



# -----------------------------------------------
# TAB: METODOLOGÍA (BPMN)
# -----------------------------------------------

with tab_flow:
    st.markdown("### ¿Cómo funciona el motor de clasificación?")
    st.divider()

    st.markdown("""
El **UNIBIM Uniclass Classifier** lee archivos IFC y asigna automáticamente
códigos **Uniclass 2015** (EF · Ss · Pr) a cada grupo de elementos constructivos,
combinando reglas deterministas con inteligencia semántica.
""")

    c1, c2, c3, c4 = st.columns(4)

    with c1:
        st.markdown("""
<div style='background:#EBF5FB;border-left:4px solid #2d6a9f;border-radius:8px;padding:1rem'>
<b style='color:#1e3a5f'>① Lectura IFC</b><br><br>
Se abre el modelo con <em>ifcopenshell</em> y se extraen todos los <code>IfcElement</code>.
Los elementos se agrupan por tipo, material y nombre para reducir trabajo redundante.
</div>
""", unsafe_allow_html=True)

    with c2:
        st.markdown("""
<div style='background:#FEF9E7;border-left:4px solid #F39C12;border-radius:8px;padding:1rem'>
<b style='color:#7D6608'>② Código EF</b><br><br>
Se detecta el dominio del modelo (estructural, HVAC, eléctrico…) y se asigna el código
<strong>EF</strong> (Entity Facets) mediante reglas basadas en el tipo IFC y el dominio.
Es siempre determinista.
</div>
""", unsafe_allow_html=True)

    with c3:
        st.markdown("""
<div style='background:#EAF2FF;border-left:4px solid #1565C0;border-radius:8px;padding:1rem'>
<b style='color:#1565C0'>③ Códigos Ss y Pr</b><br><br>
Para cada grupo se construye una frase descriptiva y se codifica con un modelo de
lenguaje especializado en construcción. Se aplican primero <strong>reglas deterministas</strong>;
si no hay coincidencia, se usa <strong>similitud coseno</strong> contra los 2 712 códigos Ss
y 8 441 códigos Pr oficiales de Uniclass 2015.
</div>
""", unsafe_allow_html=True)

    with c4:
        st.markdown("""
<div style='background:#FEF9E7;border-left:4px solid #E67E22;border-radius:8px;padding:1rem'>
<b style='color:#784212'>④ Confianza</b><br><br>
La confianza final es <code>min(conf_Ss, conf_Pr)</code>.
Las reglas devuelven <strong>0.95</strong>; el embedding devuelve el score coseno real.
Grupos con confianza &lt; 75 % se marcan <strong>⚠️ Revisar</strong> para validación manual.
</div>
""", unsafe_allow_html=True)

    st.divider()
    st.markdown("""
| Capa | Tabla Uniclass | Método principal | Fallback |
|------|---------------|-----------------|---------|
| **EF** — Entity Facets | 231 códigos | Reglas IFC type + dominio | — |
| **Ss** — Systems | 2 712 códigos | Reglas deterministas | Embedding coseno |
| **Pr** — Products | 8 441 códigos | Reglas deterministas | Embedding coseno |
""")