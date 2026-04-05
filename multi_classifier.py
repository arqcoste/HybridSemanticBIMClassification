# -----------------------------------
# MULTI-MODEL CLASSIFIER
# Procesa varios IFC y genera un CSV unificado
# -----------------------------------

import re
import os
import csv
from collections import Counter
import ifcopenshell
import openpyxl as _opxl

from engine.semantic.group_phrases import group_elements
from engine.semantic.embedding_classifier import EmbeddingClassifier

from engine.rules.domain_classifier import detect_domain_from_ifc
from engine.rules.ef_classifier import get_ef_candidates
from engine.rules.ss_classifier import classify_ss
from engine.rules.pr_classifier import classify_pr

from engine.semantic.semantic_translator import (
    build_phrase_ss,
    build_phrase_pr
)

# -------------------------
# 1. IFC FILES
# -------------------------

IFC_FILES = [
    r"C:\Users\USER\OneDrive\Documentos\Maestría\IFC PRACTICAS\WBZ-UNIBIM-XXX-IFC-EST-003.ifc",
    r"C:\Users\USER\OneDrive\Documentos\Maestría\M11\MEP\WZB-UNIBIM-XXX-IFC-SAN-001.ifc",
    r"C:\Users\USER\OneDrive\Documentos\Maestría\M11\MEP\WZB-UNIBIM-XXX-IFC-PCI-001.ifc",
    r"C:\Users\USER\OneDrive\Documentos\Maestría\M11\MEP\WZB-UNIBIM-XXX-IFC-HVAC-001.ifc",
    r"C:\Users\USER\OneDrive\Documentos\Maestría\M11\MEP\WZB-UNIBIM-XXX-IFC-ELE-001.ifc",
    r"C:\Users\USER\OneDrive\Documentos\Maestría\M11\ARQ\WZB-UNIBIM-XXX-IFC-ARQ-001.ifc",
]


def extract_model_name(path):
    """Extrae la disciplina del nombre de archivo: IFC-ARQ-001 → ARQ"""
    fname = os.path.basename(path)
    match = re.search(r"IFC-([A-Z]+)-\d+", fname, re.IGNORECASE)
    return match.group(1).upper() if match else fname.split(".")[0]


# Mapa nombre → disciplina técnica (usada como contexto en clasificadores)
DISCIPLINE_MAP = {
    "ARQ": "architecture",
    "HVAC": "hvac",
    "PCI": "fire",
    "SAN": "sanitary",
    "EST": "structural",
    "ELE": "electrical",
}


# -------------------------
# 2. EF TEXT LOOKUP
# -------------------------

print("Cargando tabla EF Uniclass...")
_wb = _opxl.load_workbook("data/uniclass/Uniclass2015_EF_v1_16.xlsx", read_only=True)
_ws = _wb.active
EF_TEXT = {}
for _row in _ws.iter_rows(values_only=True):
    code, title = str(_row[0] or "").strip(), str(_row[6] or "").strip()
    if code.startswith("EF_") and title:
        EF_TEXT[code.lower()] = title
_wb.close()

# SS: lookup desde tabla oficial Uniclass
print("Cargando tabla SS Uniclass...")
_wb2 = _opxl.load_workbook("data/uniclass/Uniclass2015_Ss_v1_40.xlsx", read_only=True)
_ws2 = _wb2.active
SS_TEXT = {}
for _row in _ws2.iter_rows(min_row=4, values_only=True):
    code, title = str(_row[0] or "").strip(), str(_row[6] or "").strip()
    if code.startswith("Ss_") and title:
        SS_TEXT[code.lower()] = title
_wb2.close()

# PR: lookup desde tabla oficial Uniclass
print("Cargando tabla PR Uniclass...")
_wb3 = _opxl.load_workbook("data/uniclass/Uniclass2015_Pr_v1_40.xlsx", read_only=True)
_ws3 = _wb3.active
PR_TEXT = {}
for _row in _ws3.iter_rows(min_row=4, values_only=True):
    code, title = str(_row[0] or "").strip(), str(_row[6] or "").strip()
    if code.startswith("Pr_") and title:
        PR_TEXT[code.lower()] = title
_wb3.close()

# -------------------------
# 3. CLASSIFIERS (cargados una sola vez)
# -------------------------

print("Cargando clasificadores SS y PR...")

classifier_ss = EmbeddingClassifier(
    model_path="models/construction_embedding_model",
    vectors_path="data/processed/ss_vectors.npy",
    metadata_path="data/processed/ss_metadata.json"
)

classifier_pr = EmbeddingClassifier(
    model_path="models/construction_embedding_model",
    vectors_path="data/processed/pr_vectors.npy",
    metadata_path="data/processed/pr_metadata.json"
)

# -------------------------
# 4. SKIP FILTER
# -------------------------

def should_skip_element(data, ef_candidates):

    if not ef_candidates or ef_candidates == ["ef_00"]:
        return True

    ifc = data.get("ifc_type", "").lower()

    IFC_IGNORE = {
        "ifcopeningelement",
        "ifcspace",
        "ifcvirtualelement",
        "ifcannotation",
        "ifcelementassembly"
    }

    return ifc in IFC_IGNORE


# -------------------------
# 5. HELPERS
# -------------------------

def _clean_name(raw):
    """Strip IFC instance ID suffix (e.g. 'BeamType:571431' → 'BeamType')."""
    if not raw or raw in ("$", ""):
        return None
    parts = raw.rsplit(":", 1)
    if len(parts) == 2 and parts[1].strip().isdigit():
        raw = parts[0].strip()
    return raw or None


# -------------------------
# 6. MAIN LOOP
# -------------------------

all_results = []

for ifc_path in IFC_FILES:

    model_name = extract_model_name(ifc_path)

    print(f"\n{'='*50}")
    print(f"MODELO: {model_name}  —  {os.path.basename(ifc_path)}")
    print(f"{'='*50}")

    try:
        model = ifcopenshell.open(ifc_path)
    except Exception as e:
        print(f"ERROR al abrir {ifc_path}: {e}")
        continue

    elements = model.by_type("IfcElement")
    print(f"Elementos encontrados: {len(elements)}")

    domain = detect_domain_from_ifc(elements)
    discipline = DISCIPLINE_MAP.get(model_name, "")

    groups = group_elements(elements)
    print(f"Grupos generados: {len(groups)}")

    # ---- PASS 1: pre-process groups, build phrases ----
    preprocessed = []
    for phrase_key, items in groups.items():
        data = dict(items[0])
        data["discipline"] = discipline
        ef_candidates = get_ef_candidates(data, domain)
        if should_skip_element(data, ef_candidates):
            continue
        ef_main = ef_candidates[0]
        preprocessed.append({
            "phrase_key": phrase_key,
            "items":      items,
            "data":       data,
            "ef_main":    ef_main,
            "phrase_ss":  build_phrase_ss(data),
            "phrase_pr":  build_phrase_pr(data),
        })

    # ---- BATCH ENCODE all unique phrases in one forward pass ----
    classifier_ss.encode_batch([g["phrase_ss"] for g in preprocessed])
    classifier_pr.encode_batch([g["phrase_pr"] for g in preprocessed])

    # ---- PASS 2: classify (all encode() calls hit cache) ----
    model_results = []
    for g in preprocessed:
        data    = g["data"]
        ef_main = g["ef_main"]
        ef_final = {"code": ef_main, "text": EF_TEXT.get(ef_main, "")}

        ss_raw   = classifier_ss.classify(g["phrase_ss"], top_k=3)
        ss_final = classify_ss(data, ss_raw)

        pr_raw   = classifier_pr.classify(g["phrase_pr"], top_k=5, ef=ef_main, data=data)
        pr_final = classify_pr(data, pr_raw)

        ss_code = ss_final["code"]
        ss_text = SS_TEXT.get(ss_code, ss_final.get("text", ""))

        pr_code = pr_final["code"]
        pr_text = PR_TEXT.get(pr_code, pr_final.get("text", ""))

        # ---- CONFIDENCE ----
        ss_conf = ss_final.get("confidence", 0.0)
        pr_conf = pr_final.get("confidence", 0.0)
        # Overall = weakest link; EF is always deterministic so not weighted
        overall_conf = min(ss_conf, pr_conf)
        confidence_pct = f"{overall_conf:.0%}"
        flag = "Revisar" if overall_conf < 0.75 else "OK"

        name_counts = Counter(
            _clean_name(item.get("name", "$")) for item in g["items"]
        )
        name_counts.pop(None, None)
        element_names = " | ".join(
            f"{n} ({c})" for n, c in sorted(name_counts.items(), key=lambda x: -x[1])
        )

        model_results.append({
            "model":         model_name,
            "domain":        domain,
            "phrase":        g["phrase_key"],
            "count":         len(g["items"]),
            "element_names": element_names,
            "ef_code":       ef_final["code"],
            "ef_text":       ef_final.get("text", ""),
            "ss_code":       ss_code,
            "ss_text":       ss_text,
            "ss_source":     ss_final.get("source", ""),
            "pr_code":       pr_code,
            "pr_text":       pr_text,
            "pr_source":     pr_final.get("source", ""),
            "confidence":    confidence_pct,
            "flag":          flag,
        })

    print(f"Grupos clasificados: {len(model_results)}")
    all_results.extend(model_results)

# -------------------------
# 6. EXPORT CSV
# -------------------------

csv_path = "classification_results_multi.csv"

with open(csv_path, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=[
        "model", "domain", "phrase", "count", "element_names",
        "ef_code", "ef_text",
        "ss_code", "ss_text", "ss_source",
        "pr_code", "pr_text", "pr_source",
        "confidence", "flag",
    ])
    writer.writeheader()
    writer.writerows(all_results)

print(f"\nTotal de filas clasificadas: {len(all_results)}")
print(f"CSV guardado en: {csv_path}")
